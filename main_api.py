import os
import re
import sys
import json
import time
from datetime import datetime
from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from supabase import create_client, Client
from dotenv import load_dotenv

# Proteção contra caracteres não-ASCII (emojis, acentos, texto vindo do Gemini) quebrando
# o stdout/stderr no Windows, cujo encoding padrão do console (cp1252) não os suporta.
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Carregar variáveis de ambiente (.env)
load_dotenv()

app = FastAPI()

# Configuração de CORS para permitir chamadas do Frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # No futuro, mude para o domínio real do seu Frontend no Vercel
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configurações do Supabase (Estas devem estar no seu arquivo .env)
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "relatorios")  # Nome do bucket no Supabase Storage

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERRO: SUPABASE_URL e SUPABASE_KEY não configurados no .env")
else:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Configurações do Portal Vale
USER_VALE = os.getenv("VALE_USER", "emanuele@sevensuprimentos.com.br")
PASS_VALE = os.getenv("VALE_PASS", "*Eas251080")

def parse_date_str(s: str):
    """Tenta vários formatos e retorna datetime.date ou None."""
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except Exception:
            continue
    return None

def log(msg: str):
    print(f"[{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}] {msg}", flush=True)

def obter_marcas_em_lote(descricoes: list, tamanho_lote: int = 60) -> list:
    """Extrai a marca do produto a partir da DESCRIÇÃO via Gemini, em lote.
    Retorna lista do mesmo tamanho/ordem de `descricoes`. Nunca propaga exceção."""
    resultado = ["Marca nao encontrada"] * len(descricoes)

    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        log("AVISO: GEMINI_API_KEY nao configurada. Pulando extracao de marca.")
        return resultado

    try:
        from google import genai
        from google.genai import types as genai_types
    except Exception as e:
        log(f"AVISO: SDK google-genai indisponivel ({e}). Pulando extracao de marca.")
        return resultado

    modelo = os.getenv("GEMINI_MODEL", "gemini-3.1-flash-lite")
    client = genai.Client(api_key=api_key)

    total = len(descricoes)
    for inicio in range(0, total, tamanho_lote):
        fim = min(inicio + tamanho_lote, total)
        lote = descricoes[inicio:fim]

        itens = "\n".join(f"{i}: {(desc or '').strip()}" for i, desc in enumerate(lote))
        prompt = (
            "Voce e um extrator de marcas de produtos industriais. Para cada item numerado "
            "abaixo, identifique a MARCA do produto (fabricante), seguindo estas regras:\n"
            "- So extraia a marca se ela estiver EXPLICITAMENTE escrita no texto. NUNCA "
            "invente ou deduza uma marca a partir do tipo de produto.\n"
            "- NUNCA retorne \"VALE\" como marca — e o nome do cliente comprador, nao do "
            "fabricante.\n"
            "- Se nao houver marca identificavel, retorne exatamente \"Marca nao encontrada\".\n"
            "- Retorne a marca em maiusculas, sem codigo de peca junto.\n"
            "- Responda APENAS com um JSON array puro, sem markdown, no formato: "
            '[{"indice": N, "marca": "..."}]\n\n'
            f"Itens:\n{itens}"
        )

        try:
            response = client.models.generate_content(
                model=modelo,
                contents=prompt,
                config=genai_types.GenerateContentConfig(
                    response_mime_type="application/json",
                    temperature=0,
                ),
            )
            dados = json.loads(response.text)
            preenchidos = 0
            for item in dados:
                idx = item.get("indice")
                marca = item.get("marca")
                if isinstance(idx, int) and 0 <= idx < len(lote) and marca:
                    resultado[inicio + idx] = str(marca).strip().upper()
                    preenchidos += 1
            log(f"Marcas: lote {inicio}-{fim - 1}: {preenchidos}/{len(lote)} preenchidas.")
        except Exception as e:
            log(f"AVISO: falha ao obter marcas do lote {inicio}-{fim - 1}: {e}")
            continue

    return resultado

def executar_robo_selenium(data_usuario: str, filename: str, pular_verificacao_duplicados: bool = False):
    """Lógica principal do robô Selenium refatorada para rodar sem interface (headless)."""

    inicio_total = time.time()
    log(f"=== ROBÔ INICIADO | data={data_usuario} | arquivo={filename} ===")
    if pular_verificacao_duplicados:
        log("MODO TESTE: verificacao de eventos duplicados DESATIVADA - eventos ja existentes serao reprocessados")

    # 1. Preparar data (Aceita 6 ou 8 dígitos: DDMMAA ou DDMMAAAA)
    if len(data_usuario) == 6:
        HOJE_str = f"{data_usuario[:2]}/{data_usuario[2:4]}/{data_usuario[4:]}"
    else:
        HOJE_str = f"{data_usuario[:2]}/{data_usuario[2:4]}/{data_usuario[4:]}" # Pega os 4 dígitos do ano se houver

    HOJE = parse_date_str(HOJE_str)
    if not HOJE:
        log(f"ERRO: Não foi possível converter a data '{data_usuario}'")
        return

    # 2. Configurar Selenium Headless para o Servidor (Railway/Docker)
    chrome_options = Options()
    chrome_options.add_argument("--headless=new") # IMPORTANTE: Modo invisível
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    
    # --- OTIMIZAÇÕES REDUÇÃO DE MEMÓRIA E CPU ---
    chrome_options.add_argument("--disable-extensions") # Não carregar extensões
    chrome_options.add_argument("--disable-software-rasterizer") # Exige menos da CPU onde já não tem GPU
    chrome_options.add_argument("--disable-logging") # Menos escrita de logs nativos do chrome
    chrome_options.add_argument("--log-level=3")     # Silencia saída desnecessária
    chrome_options.add_argument("--blink-settings=imagesEnabled=false") # 🚀 DESATIVA IMAGENS (salva MUITA RAM e Banda)
    chrome_options.page_load_strategy = 'eager' # 🚀 CARREGA MAIS RÁPIDO (não espera recursos inúteis da página, só o DOM)
    
    # Localiza o binário do Chrome no Railway
    chrome_bin = os.getenv("CHROME_BIN")
    if chrome_bin:
        chrome_options.binary_location = chrome_bin
    
    # Inicializa o driver apenas UMA vez com as opções corretas
    driver = webdriver.Chrome(options=chrome_options)
    
    EXCEL_PATH = f"/tmp/{filename}"
    ESTADOS = ['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO']

    try:
        # Prepara Planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Eventos"
        ws.append(["Numero do evento", "Titulo", "UF(VALE)", "DATA", "DESCRIÇÃO", "QTDE", "UNID. MED", "pagina de descrição", "Marca"])
        
        wait = WebDriverWait(driver, 20)
        log("Acessando página de login...")
        driver.get("https://vale.coupahost.com/sessions/supplier_login")

        # Login
        wait.until(EC.presence_of_element_located((By.ID, "user_login")))
        driver.find_element(By.ID, "user_login").send_keys(USER_VALE)
        driver.find_element(By.ID, "user_password").send_keys(PASS_VALE, Keys.RETURN)
        log("Login enviado. Aguardando carregamento...")

        # Filtro de data (mesma lógica)
        try:
            time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
            time_filter.click()
            time.sleep(5)
            time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
            time_filter.click()
            log("Filtro de data aplicado.")
        except:
            log("AVISO: Filtro de data não encontrado, continuando sem ele.")

        # Coleta de dados - Parte 1: Listagem de Eventos
        log(f"--- FASE 1: Coletando lista de eventos para {HOJE_str} ---")
        encontrou_ontem = False
        pagina = 1
        total_coletados = 0
        while True:
            log(f"Pagina {pagina}: aguardando tabela de eventos...")
            time.sleep(5)
            try:
                tbody = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="quote_request_table_tag"]')))
                # Usar index para evitar stale element reference
                num_linhas = len(tbody.find_elements(By.TAG_NAME, "tr"))
                log(f"Pagina {pagina}: {num_linhas} linhas encontradas na tabela.")

                for i in range(num_linhas):
                    try:
                        # Re-fetch as linhas a cada iteração para evitar stale
                        tbody = driver.find_element(By.XPATH, '//*[@id="quote_request_table_tag"]')
                        linhas = tbody.find_elements(By.TAG_NAME, "tr")
                        if i >= len(linhas): break
                        linha = linhas[i]

                        colunas = linha.find_elements(By.TAG_NAME, "td")
                        if not colunas or len(colunas) < 7: continue

                        yellow_flags = linha.find_elements(By.CSS_SELECTOR, "img[src*='flag_yellow']")
                        if yellow_flags: continue

                        status_text = colunas[4].text.strip()
                        if "Concluído" in status_text: continue

                        data_inicio_str = colunas[2].text.strip()
                        data_inicio = parse_date_str(data_inicio_str)
                        if data_inicio is None: continue

                        if data_inicio < HOJE:
                            encontrou_ontem = True
                            log(f"Data anterior encontrada ({data_inicio}), encerrando coleta.")
                            break

                        if data_inicio != HOJE: continue

                        numero_evento = colunas[0].find_element(By.TAG_NAME, "a").text.strip()

                        # Verifica se o evento já existe no banco de dados
                        if not pular_verificacao_duplicados:
                            try:
                                res_db = supabase.table("eventos_coletados").select("id").eq("numero_evento", numero_evento).execute()
                                if res_db.data:
                                    log(f"AVISO: Evento {numero_evento} ja existe no banco. Pulando.")
                                    continue
                            except Exception as e_db:
                                log(f"ERRO ao verificar evento {numero_evento} no banco: {e_db}")

                        data_final = colunas[3].text.strip()
                        total_coletados += 1
                        log(f"  [{total_coletados}] Evento coletado: {numero_evento} | prazo: {data_final}")
                        ws.append([numero_evento, '', '', data_final, '', '', '', '', ''])
                    except Exception as e:
                        log(f"ERRO na linha {i} da pagina {pagina}: {e}")
                        continue
            except Exception as e:
                log(f"ERRO ao acessar tabela na pagina {pagina}: {e}")
                break

            if encontrou_ontem: break
            try:
                proximo = driver.find_element(By.CLASS_NAME, "next_page")
                driver.execute_script("arguments[0].click();", proximo)
                pagina += 1
                time.sleep(3)
            except:
                log(f"Sem proxima pagina. Total de paginas percorridas: {pagina}.")
                break

        wb.save(EXCEL_PATH)
        log(f"Fase 1 concluida: {total_coletados} evento(s) coletado(s). Iniciando detalhamento...")

        # --- DETALHA CADA EVENTO ---
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Eventos"]

        todos_eventos = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
        total_eventos = len(todos_eventos)
        log(f"--- FASE 2: Detalhando {total_eventos} evento(s) ---")

        for idx_ev, row in enumerate(ws.iter_rows(min_row=2), start=1):
            evento = row[0].value
            if not evento:
                continue

            log(f"[{idx_ev}/{total_eventos}] Detalhando evento: {evento}")

            driver.get(f"https://vale.coupahost.com/quotes/external_responses/{evento}/edit")
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

            # --- VERIFICA EXISTÊNCIA DA PÁGINA DE DESCRIÇÃO ---
            try:
                botoes1 = driver.find_elements(By.XPATH, '//*[@id="pageContentWrapper"]/div[3]/div[2]/a[2]/span')
                if not botoes1:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    botoes2 = driver.find_elements(By.ID, 'quote_response_submit')
                    if botoes2:
                        log(f"  Evento {evento}: botao de submit encontrado, clicando...")
                        botoes2[0].click()
            except Exception:
                log(f"  ERRO ao verificar pagina de descricao do evento {evento}")
                row[7].value = "Erro ao verificar página de descrição"

            # Scroll e abre seção das informações
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            # Tenta encontrar o seletor principal ou o fallback conforme solicitado
            try:
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "s-expandLines")))
                seletor_atual = (By.CLASS_NAME, "s-expandLines")
                query_css = ".s-expandLines"
            except:
                log("⚠️ s-expandLines não encontrado, tentando fallback...")
                fallback_css = ".sidebar.-supplier.-borderLeft.flexPosition__element.-shrink.s-expandSidebar.-clickable"
                try:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, fallback_css)))
                    seletor_atual = (By.CSS_SELECTOR, fallback_css)
                    query_css = fallback_css
                except:
                    log(f"⚠️ Nenhum seletor de expansão encontrado no evento {evento}")
                    continue

            elementos = driver.find_elements(*seletor_atual)

            if not elementos:
                log(f"⚠️ Nenhum elemento de expansão encontrado no evento {evento}")
                continue

            # Duplicar a linha do evento pelo número de elementos encontrados
            linhas_evento = [row]
            if len(elementos) > 1:
                for i in range(len(elementos) - 1):
                    nova_linha = [evento, row[1].value, row[2].value, row[3].value, '', '', '', '', '']
                    ws.append(nova_linha)
                wb.save(EXCEL_PATH)
                linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]

            # Percorre cada s-expandLines e coleta os dados (re-fetch a cada iteração, marca processed via JS)
            def click_element_retry(el, attempts=4, pause=0.4):
                from selenium.common.exceptions import (
                    StaleElementReferenceException,
                    ElementClickInterceptedException,
                    ElementNotInteractableException,
                    WebDriverException,
                )
                for _ in range(attempts):
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.15)
                        el.click()
                        return True
                    except (StaleElementReferenceException, ElementClickInterceptedException, ElementNotInteractableException, WebDriverException):
                        try:
                            driver.execute_script("arguments[0].click();", el)
                            return True
                        except Exception:
                            time.sleep(pause)
                return False

            # determina quantos existem no DOM no momento
            total = driver.execute_script(f"return document.querySelectorAll('{query_css}').length")
            if total == 0:
                log(f"⚠️ Nenhum elemento de expansão encontrado no evento {evento}")
                continue

            # duplicar linha já feito acima; garante linhas_evento atualizado
            linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]

            processed = 0
            idx = 0
            while processed < total and idx < total:
                # re-obtem a lista sempre
                try:
                    elementos = driver.find_elements(*seletor_atual)
                except Exception:
                    time.sleep(0.3)
                    elementos = driver.find_elements(*seletor_atual)

                if idx >= len(elementos):
                    # DOM encolheu — tenta refetch algumas vezes
                    retry_try = 0
                    while retry_try < 3 and idx >= len(elementos):
                        time.sleep(0.4)
                        elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")
                        retry_try += 1
                    if idx >= len(elementos):
                        log(f"⚠️ Índice {idx} fora do range atual ({len(elementos)}). Pulando.")
                        idx += 1
                        continue

                el = elementos[idx]

                # evita re-processar elemento já marcado
                try:
                    already = el.get_attribute('data-processed')
                except:
                    already = None

                if already:
                    idx += 1
                    processed += 1
                    continue

                # tenta clicar de forma robusta
                if not click_element_retry(el, attempts=4, pause=0.4):
                    log(f"⚠️ Falha ao clicar no expandLines index {idx} do evento {evento}")
                    # marca como processado para não travar loop
                    try:
                        driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                    except Exception:
                        pass
                    idx += 1
                    processed += 1
                    continue

                # após clique, espera conteúdo de detalhe carregar (linha expandida)
                try:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".s-itemsAndServicesLine.-expanded")))
                    time.sleep(0.25)
                except Exception:
                    time.sleep(0.4)

                # atualiza linhas_evento porque podem ter sido adicionadas
                linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]
                try:
                    linha_atual = linhas_evento[idx]
                except Exception:
                    # se não existir, tenta mapear para próxima disponível
                    if linhas_evento:
                        linha_atual = linhas_evento[-1]
                    else:
                        log(f"⚠️ Não há linha disponível para evento {evento} no idx {idx}")
                        # marca e segue
                        try:
                            driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                        except Exception:
                            pass
                        idx += 1
                        processed += 1
                        continue

                # coleta campos — seletores por classe CSS, compatíveis com ambos layouts
                log(f"  Item {idx + 1}/{total}: coletando campos...")

                # Escopo: a linha atualmente expandida
                try:
                    wait.until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, ".s-itemsAndServicesLine.-expanded")
                    ))
                    linha_dom = driver.find_element(By.CSS_SELECTOR, ".s-itemsAndServicesLine.-expanded")
                except Exception:
                    time.sleep(1)
                    try:
                        linha_dom = driver.find_element(By.CSS_SELECTOR, ".s-itemsAndServicesLine.-expanded")
                    except Exception:
                        linha_dom = driver

                def buscar_texto(seletores, escopo=None):
                    """Tenta uma lista de seletores CSS e retorna o primeiro texto não vazio."""
                    alvo = escopo if escopo is not None else linha_dom
                    for sel in seletores:
                        try:
                            el = alvo.find_element(By.CSS_SELECTOR, sel)
                            txt = (el.text or "").strip()
                            if txt:
                                return txt
                        except Exception:
                            continue
                    return ""

                # --- QUANTIDADE ---
                qtde = buscar_texto([".s-quantity .s-value"])
                linha_atual[5].value = qtde if qtde else 'N/A'

                # --- UNIDADE ---
                unid = buscar_texto([".s-quantity .s-unit"])
                linha_atual[6].value = unid if unid else 'N/A'

                # --- DESCRIÇÃO (cascata: campo padrão -> custom field do Lance spot) ---
                descri = buscar_texto([
                    ".s-extended_description p.s-textField",
                    ".s-request_line_custom_fields dd.itemReview span",
                    ".s-request_line_custom_fields dd span",
                ])
                if descri:
                    desejado = re.search(r'PT\s*\|\|\s*(.*?)\*{3,}', descri, re.DOTALL)
                    linha_atual[4].value = desejado.group(1).strip() if desejado else descri
                else:
                    linha_atual[4].value = 'N/A'
                    log(f"  AVISO: descricao nao encontrada no evento {evento} item {idx+1}")

                # --- TÍTULO (aceita separador "||" ou "|") ---
                titulo_raw = buscar_texto([".s-description p.s-textField"])
                if titulo_raw:
                    m_tit = re.match(r"^\s*(\d+)", titulo_raw)
                    linha_atual[1].value = m_tit.group(1) if m_tit else titulo_raw.split("|")[0].strip()
                else:
                    linha_atual[1].value = "Titulo nao encontrado"

                # --- UF (cascata: endereço de entrega -> varredura geral da linha) ---
                found = None
                endereco_txt = buscar_texto([
                    ".s-ship_to_address .s-itemsAndServicesAddressLine",
                    ".s-ship_to_address .addressLines",
                ])

                if endereco_txt:
                    texto = endereco_txt.upper()
                    # padrão "CEP CIDADE UF"
                    m_uf = re.search(r'\d{5}-?\d{3}\s+.+?\s+([A-Z]{2})\s*(?:\n|$)', texto)
                    if m_uf and m_uf.group(1) in ESTADOS:
                        found = m_uf.group(1)
                    if not found:
                        m_uf = re.search(r'-\s*([A-Z]{2})\s*-\s*BR', texto)
                        if m_uf and m_uf.group(1) in ESTADOS:
                            found = m_uf.group(1)
                    if not found:
                        for t in re.findall(r'\b[A-Z]{2}\b', texto):
                            if t in ESTADOS:
                                found = t
                                break

                # fallback: varre todo o texto da linha expandida
                if not found:
                    try:
                        combined = (linha_dom.text or "").upper()
                        m_uf = re.search(r'-\s*([A-Z]{2})\s*-\s*BR', combined)
                        if m_uf and m_uf.group(1) in ESTADOS:
                            found = m_uf.group(1)
                        if not found:
                            for sig in ESTADOS:
                                if re.search(r'\b' + re.escape(sig) + r'\b', combined):
                                    found = sig
                                    break
                    except Exception:
                        pass

                linha_atual[2].value = found if found else 'UF nao encontrada'
                if not found:
                    log(f"  AVISO: UF nao encontrada no evento {evento} item {idx+1}")

                log(f"  Item {idx + 1}/{total}: titulo={linha_atual[1].value!r} | uf={linha_atual[2].value!r} | qtde={linha_atual[5].value!r} | unid={linha_atual[6].value!r}")

                # fecha o detalhe (tenta vários métodos)
                try:
                    time.sleep(0.2)
                    fechar = None
                    try:
                        fechar = driver.find_element(By.CSS_SELECTOR, "button.button.s-cancel")
                    except Exception:
                        try:
                            fechar = driver.find_element(By.XPATH, "//button[contains(concat(' ', normalize-space(@class), ' '), ' s-cancel ') and contains(., 'Cancelar')]")
                        except Exception:
                            fechar = None
                    if fechar:
                        click_element_retry(fechar, attempts=3, pause=0.2)
                        time.sleep(0.25)
                except Exception:
                    pass

                # marca como processado (para não reprocessar se DOM reorganizar)
                try:
                    driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                except Exception:
                    pass

                processed += 1
                idx += 1

            wb.save(EXCEL_PATH)
            log(f"[{idx_ev}/{total_eventos}] Evento {evento} finalizado.")

        log("--- FASE 3: Ordenando planilha e salvando no banco ---")
        # Ordena a planilha por "Numero do evento" (coluna A) para agrupar linhas com o mesmo número
        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb["Eventos"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            descricoes = [r[4] if len(r) > 4 else '' for r in rows]
            log(f"--- Extraindo marcas via Gemini para {len(descricoes)} item(ns) ---")
            marcas = obter_marcas_em_lote(descricoes)
            rows = [r[:8] + (marcas[i],) for i, r in enumerate(rows)]

            def sort_key(row):
                v = row[0]
                if v is None:
                    return (1, "")
                s = str(v).strip()
                try:
                    return (0, int(s))      # números antes de strings, ordenados numericamente
                except Exception:
                    return (1, s.lower())   # strings ordenadas alfabeticamente

            rows_sorted = sorted(rows, key=sort_key)

            # remove linhas antigas (todas a partir da linha 2) e escreve ordenado
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            
            eventos_para_inserir = []
            vistos_db = set()

            for r in rows_sorted:
                ws.append(list(r))
                ne = str(r[0]).strip()
                titulo = str(r[1]) if len(r) > 1 and r[1] is not None else ""
                chave = (ne, titulo)
                if ne and ne != "None" and chave not in vistos_db:
                    vistos_db.add(chave)
                    eventos_para_inserir.append({
                        "numero_evento": ne,
                        "titulo": titulo,
                        "uf": str(r[2]) if r[2] is not None else "",
                        "data_evento": str(r[3]) if r[3] is not None else "",
                        "descricao": str(r[4]) if r[4] is not None else "",
                        "quantidade": str(r[5]) if r[5] is not None else "",
                        "unidade": str(r[6]) if r[6] is not None else "",
                        "marca": str(r[8]) if len(r) > 8 and r[8] is not None else ""
                    })

            wb.save(EXCEL_PATH)

            # Salva os novos eventos no banco de dados para evitar futuras duplicidades
            if eventos_para_inserir:
                try:
                    supabase.table("eventos_coletados").upsert(eventos_para_inserir, on_conflict="numero_evento,titulo").execute()
                    log(f"Banco atualizado: {len(eventos_para_inserir)} evento(s) inserido(s)/atualizados.")
                except Exception as e_db:
                    log(f"ERRO ao inserir eventos no banco: {e_db}")
            else:
                log("Nenhum evento novo para inserir no banco.")

        except Exception as e:
            log(f"ERRO ao ordenar/salvar planilha ou banco: {e}")

        log("--- FASE 4: Fazendo upload para o Supabase Storage ---")
        with open(EXCEL_PATH, "rb") as f:
            supabase.storage.from_(SUPABASE_BUCKET).upload(
                path=filename,
                file=f,
                file_options={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
            )

        res = supabase.storage.from_(SUPABASE_BUCKET).get_public_url(filename)
        log(f"Upload concluido. URL publica: {res}")

        if os.path.exists(EXCEL_PATH):
            os.remove(EXCEL_PATH)

    except Exception as e:
        log(f"ERRO FATAL no robô: {e}")
        log(f"Arquivo Excel preservado para recuperacao manual em: {EXCEL_PATH}")
    finally:
        driver.quit()
        elapsed = time.time() - inicio_total
        log(f"=== ROBÔ FINALIZADO | tempo total: {elapsed:.1f}s ===")

@app.get("/")
def read_root():
    return {"status": "Backend Online", "projeto": "Seven Suprimentos - Automação Vale"}

@app.post("/run-robot")
def run_robot(data: str, background_tasks: BackgroundTasks, pular_verificacao_duplicados: bool = False):
    """Inicia o robô como uma tarefa de fundo."""
    if not data or len(data) not in [6, 8]:
        raise HTTPException(status_code=400, detail="Data deve estar no formato DDMMAA ou DDMMAAAA")

    # Padroniza para DDMMAA para gerar o nome do arquivo, mas passa a data completa
    data_formatada = data if len(data) == 6 else f"{data[:4]}{data[6:]}"
    filename = f"eventos_{data_formatada}_{int(time.time())}.xlsx"

    background_tasks.add_task(executar_robo_selenium, data, filename, pular_verificacao_duplicados)

    return {
        "status": "Iniciado",
        "message": f"Robô iniciado para a data {data}. O arquivo será enviado para o Supabase Storage em alguns minutos.",
        "filename": filename,
        "modo_teste_duplicados": pular_verificacao_duplicados
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
